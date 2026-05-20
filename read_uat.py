import json
import openpyxl

path = r"C:\Users\zulub\Desktop\STELLA LUMEN\SORTING\RIGEL 4.1.0\Rigel Testing Grid - Suppliers, Investments, Loans.xlsx"
wb = openpyxl.load_workbook(path)
all_data = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    all_data[sheet_name] = rows

with open("uat_output.json", "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print("Done - wrote uat_output.json")
